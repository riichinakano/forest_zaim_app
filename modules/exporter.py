"""
エクスポートモジュール

データのCSV/Excel出力を担当します。
"""

from io import BytesIO
from typing import Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


def export_to_csv(df: pd.DataFrame, filename: Optional[str] = None) -> bytes:
    """
    DataFrameをCSV形式でエクスポート

    Args:
        df: エクスポートするDataFrame
        filename: ファイル名（未使用、互換性のため保持）

    Returns:
        bytes: CSV data (UTF-8 BOM付き、Excel対応)

    Examples:
        >>> csv_data = export_to_csv(df)
        >>> with open('output.csv', 'wb') as f:
        ...     f.write(csv_data)
    """
    # UTF-8 BOM付きで出力（Excelでの文字化け防止）
    csv_string = df.to_csv(index=False, encoding='utf-8-sig')
    return csv_string.encode('utf-8-sig')


def export_to_excel(
    df: pd.DataFrame,
    filename: Optional[str] = None,
    sheet_name: str = "月次推移"
) -> bytes:
    """
    DataFrameをExcel形式でエクスポート（書式設定付き）

    Args:
        df: エクスポートするDataFrame
        filename: ファイル名（未使用、互換性のため保持）
        sheet_name: シート名

    Returns:
        bytes: Excel data

    Examples:
        >>> excel_data = export_to_excel(df, sheet_name="売上推移")
        >>> with open('output.xlsx', 'wb') as f:
        ...     f.write(excel_data)
    """
    # Workbook作成
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # DataFrameを行データに変換
    rows = dataframe_to_rows(df, index=False, header=True)

    # データを書き込み
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            # ヘッダー行のスタイル
            if r_idx == 1:
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")

            # 罫線
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = thin_border

    # 列幅の調整
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column_letter].width = adjusted_width

    # 数値列の書式設定
    for col_idx, col_name in enumerate(df.columns, 1):
        # 金額列（4月〜3月、年間合計）
        if col_name in ['4月', '5月', '6月', '7月', '8月', '9月',
                       '10月', '11月', '12月', '1月', '2月', '3月', '年間合計']:
            for row_idx in range(2, len(df) + 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.number_format = '#,##0'  # 3桁カンマ区切り

        # 前年比列
        elif col_name == '前年比':
            for row_idx in range(2, len(df) + 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = cell.value

                # パーセント表示
                if value is not None and not pd.isna(value):
                    cell.number_format = '0.0"%"'

                    # 色分け（増加=緑、減少=赤）
                    if isinstance(value, (int, float)):
                        if value > 0:
                            cell.font = Font(color="00B050")  # 緑
                        elif value < 0:
                            cell.font = Font(color="FF0000")  # 赤

    # 年間合計列を太字に
    if '年間合計' in df.columns:
        total_col_idx = df.columns.get_loc('年間合計') + 1
        for row_idx in range(1, len(df) + 2):
            cell = ws.cell(row=row_idx, column=total_col_idx)
            cell.font = Font(bold=True, size=11)

    # BytesIOに保存
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()


def create_download_filename(account_name: str, file_type: str) -> str:
    """
    ダウンロード用ファイル名を生成

    Args:
        account_name: 科目名
        file_type: ファイルタイプ ('csv' または 'excel')

    Returns:
        str: ファイル名

    Examples:
        >>> create_download_filename('売上高', 'csv')
        '売上高_月次推移.csv'
        >>> create_download_filename('役員報酬', 'excel')
        '役員報酬_月次推移.xlsx'
    """
    extension = 'xlsx' if file_type == 'excel' else 'csv'
    # ファイル名に使えない文字を置換
    safe_name = account_name.replace('/', '_').replace('\\', '_')
    return f"{safe_name}_月次推移.{extension}"
